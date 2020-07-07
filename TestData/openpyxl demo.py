import openpyxl


book = openpyxl.load_workbook("C:\\Users\\Sameer\\Desktop\\Pydoc.xlsx")
sheet = book.active
cell =  sheet.cell(row=3,column=1)
# print(cell.value)
sheet.cell(row=2,column=2).value ='Rahul'

# print(sheet['B6'].value)


for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):

        print(sheet.cell(row=i,column=j).value)
